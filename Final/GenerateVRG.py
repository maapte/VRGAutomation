"""
__author__ : Karthik
__maintainer__ : Manish  & karthik
__organization__ : Deloitte
"""

from datetime import date
from datetime import datetime

import openpyxl
from flask import json
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, Border, Side
from StandalonePage import StandalonePage
from Steps import Steps
from constant.ApplicationConstant import *

from constant.TestCaseConstant import (
    TEST_CASE_ERROR_MESSAGE,
    TEST_CASE_ERROR_NAME,
    TEST_CASE_ERROR_DESCRIPTION,
    TEST_CASE_SUBJECT,
    TEST_CASE_OBJECTIVE,
)


class GenerateVrg:
    def get_site_name_code(self, site_name):
        if str(site_name) == "banking-cibc" or str(site_name) == "banking-simplii":
            return str("olb")
        elif str(site_name) == "personal-banking":
            return str("pb")
        else:
            return str(site_name)

    def get_page_accessibility_message(self, page_accessibility):
        if str(page_accessibility) == "true" or str(page_accessibility) == "{dynamic}":
            return str("Page accessibility available ")
        else:
            return str("Page accessibility not available ")

    def get_page_hiearchy(self, page_hierarchy, test_case=""):
        if len(page_hierarchy) > 1:
            if test_case == "test":
                if page_hierarchy == "{dynamic}":
                    return page_hierarchy
                return "-".join(list(page_hierarchy[: len(page_hierarchy) - 1]))
            return ">".join(list(page_hierarchy[: len(page_hierarchy) - 1]))
        return "".join(list(page_hierarchy))

    def get_page_name(self, page_name):
        return ">".join(list(page_name[len(page_name) - 1 :]))

    def get_otvc_generation(self, raw_obj, data):
        temp_lst = []
        temp_step_lst = []
        temp_step = []
        temp_form_obj = []

        for i in range(0, len(raw_obj.pages)):
            temp_lst.insert(i, raw_obj.pages[i])

        temp_count = 0
        for i in range(0, len(temp_lst)):
            if temp_lst[i]["is_otvc"] == "true":
                temp_count = temp_count + 2
            temp_count = temp_count + 1

        for i in range(0, temp_count):
            if temp_lst[i]["is_otvc"] == "true":
                standalone_page_detail = StandalonePage()
                if data['application_type'] == APPLICATION_EMBER.lower():
                    page_hierarchy_str = temp_lst[i]["page_hierarchy"]
                    page_name_str = temp_lst[i]["page_name"]
                    page_name_str = str(page_name_str)
                    page_hierarchy_str = str(page_hierarchy_str)
                    standalone_page_detail.set_page_path(page_name_str)
                    standalone_page_detail.set_page_hierarchy(page_hierarchy_str)
                    standalone_page_detail.set_page_name(page_name_str)
                if data['application_type'] == APPLICATION_ANGULAR.lower():
                    page_name_str = temp_lst[i]["page_name"]
                    page_name_str = str(page_name_str)
                    standalone_page_detail.set_page_name(page_name_str)
                    page_hierarchy_str = temp_lst[i]["page_hierarchy"]
                    standalone_page_detail.set_page_path(page_name_str)
                    standalone_page_detail.set_page_hierarchy(str(page_hierarchy_str))

                standalone_page_detail.set_page_referrer(temp_lst[i]["page_referrer"])
                if len(temp_lst[i]["user_id"]) == 0:
                    standalone_page_detail.set_user_id("false")
                else:
                    standalone_page_detail.set_user_id("true")
                if temp_lst[i]["event_error"] == "false":
                    standalone_page_detail.set_error_code("")
                    standalone_page_detail.set_event_error("false")
                if temp_lst[i]["event_error"] == "true":
                    standalone_page_detail.set_event_error("true")
                    standalone_page_detail.set_error_message("{dynamic}")
                    standalone_page_detail.set_error_code("{dynamic}")
                    standalone_page_detail.set_error_type("{dynamic}")
                    standalone_page_detail.set_error_sub_type("{dynamic}")
                    standalone_page_detail.set_error_field("{dynamic}")
                if temp_lst[i]["user_auth_state"] == "{authenticated}":
                    standalone_page_detail.set_user_auth_state("{authenticated}")
                if temp_lst[i]["user_auth_state"] == "{not-authenticated}":
                    standalone_page_detail.set_user_auth_state("{not-authenticated}")
                if (
                    temp_lst[i]["user_auth_state"]
                    == "{not-authenticated | authenticated}"
                ):
                    standalone_page_detail.set_user_auth_state(
                        "{not-authenticated | authenticated}"
                    )

                if temp_lst[i]["advertising"] == "true":
                    standalone_page_detail.set_advertising("true")
                    standalone_page_detail.set_events_advertising("true")
                    standalone_page_detail.set_events_ad_click("true")
                    standalone_page_detail.set_advertising_tracking_code("{dynamic}")
                    standalone_page_detail.set_advertising_location("{dynamic}")
                    standalone_page_detail.set_advertising_type("{dynamic}")
                if temp_lst[i]["is_login"] == "true":
                    standalone_page_detail.set_is_login("true")
                if temp_lst[i]["is_otvc"] == "true":
                    standalone_page_detail.set_is_otvc("true")

                standalone_prompt_obj = standalone_page_detail
                standalone_prompt_obj.set_is_otvc("")
                str_temp = standalone_prompt_obj.get_page_name + " (OTVC PROMPT)"
                standalone_prompt_obj.set_page_name(str(str_temp))
                standalone_prompt_obj.set_is_otvc_success("")
                standalone_prompt_obj.set_is_otvc_prompt("true")
                standalone_prompt_obj.set_is_login("")
                standalone_prompt_obj.set_page_view("false")
                standalone_prompt_obj.set_user_id("{dynamic}")
                standalone_prompt_obj.set_user_type("{dynamic}")
                page = json.dumps(standalone_prompt_obj.__dict__)
                raw_val = json.loads(page)
                temp_lst.insert(i + 1, raw_val)

                standalone_success_obj = standalone_page_detail
                standalone_success_obj.set_is_otvc("")
                str_temp = standalone_success_obj.get_page_name
                str_temp = str_temp.replace(" (OTVC PROMPT)", " (OTVC SUCCESS)")
                standalone_success_obj.set_page_name(str(str_temp))
                standalone_success_obj.set_is_otvc_success("true")
                standalone_success_obj.set_is_otvc_prompt("")
                standalone_success_obj.set_is_login("")
                standalone_success_obj.set_page_view("false")
                standalone_success_obj.set_user_id("{dynamic}")
                standalone_success_obj.set_user_type("{dynamic}")
                page = json.dumps(standalone_success_obj.__dict__)
                raw_val = json.loads(page)
                temp_lst.insert(i + 2, raw_val)

        raw_obj.pages = temp_lst

        # step List
        for i in range(0, len(raw_obj.forms)):
            for j in range(0, len(raw_obj.forms[i]["steps"])):
                temp_step_lst.append(raw_obj.forms[i]["steps"][j])

        temp_step_count = 0
        for i in range(0, len(raw_obj.forms)):
            for j in range(0, len(raw_obj.forms[i]["steps"])):
                temp_step_count = temp_step_count + 1

        form_obj = ""
        for i in range(0, len(raw_obj.forms)):
            for j in range(0, len(raw_obj.forms[i]["steps"])):
                if raw_obj.forms[i]["steps"][j]["is_otvc"] == "true":
                    steps_obj = Steps()

                    if data['application_type'] == APPLICATION_EMBER.lower():
                        page_hierarchy_steps_str = raw_obj.forms[i]["steps"][j]["page_hierarchy"]
                        page_name_steps_str = raw_obj.forms[i]["steps"][j]["page_name"]
                        page_name_steps_str = str(page_name_steps_str)
                        page_hierarchy_steps_str = str(page_hierarchy_steps_str)
                        steps_obj.set_page_hierarchy(page_hierarchy_steps_str)
                        steps_obj.set_page_name(page_name_steps_str)
                        steps_obj.set_page_path(page_hierarchy_steps_str)
                    if data['application_type'] == APPLICATION_ANGULAR.lower():
                        page_name_steps_str = raw_obj.forms[i]["steps"][j]["page_name"]
                        page_hierarchy_steps_str = raw_obj.forms[i]["steps"][j]["page_hierarchy"]
                        page_name_steps_str = str(page_name_steps_str)
                        page_hierarchy_steps_str = str(page_hierarchy_steps_str)
                        steps_obj.set_page_hierarchy(page_hierarchy_steps_str)
                        steps_obj.set_page_path(page_name_steps_str)
                        steps_obj.set_page_hierarchy(page_hierarchy_steps_str)

                    steps_obj.set_step_name(raw_obj.forms[i]["steps"][j]["step_name"])
                    steps_obj.set_user_id(raw_obj.forms[i]["steps"][j]["user_id"])
                    steps_obj.set_user_type(raw_obj.forms[i]["steps"][j]["user_type"])
                    steps_obj.set_user_auth_state(
                        raw_obj.forms[i]["steps"][j]["user_auth_state"]
                    )
                    steps_obj.set_from_transaction(
                        raw_obj.forms[i]["steps"][j]["from_transaction"]
                    )
                    steps_obj.set_to_transaction(
                        raw_obj.forms[i]["steps"][j]["to_transaction"]
                    )
                    steps_obj.set_is_external(
                        raw_obj.forms[i]["steps"][j]["is_external"]
                    )
                    steps_obj.set_form_view(raw_obj.forms[i]["steps"][j]["form_view"])
                    steps_obj.set_form_qualify(
                        raw_obj.forms[i]["steps"][j]["form_qualify"]
                    )
                    steps_obj.set_form_submit(
                        raw_obj.forms[i]["steps"][j]["form_submit"]
                    )
                    steps_obj.set_product_id(raw_obj.forms[i]["steps"][j]["product_id"])
                    steps_obj.set_parent_product(
                        raw_obj.forms[i]["steps"][j]["parent_product"]
                    )
                    steps_obj.set_adjudication(
                        raw_obj.forms[i]["steps"][j]["adjudication"]
                    )
                    steps_obj.set_product_positioning(
                        raw_obj.forms[i]["steps"][j]["product_positioning"]
                    )
                    steps_obj.set_product_grouping(
                        raw_obj.forms[i]["steps"][j]["product_grouping"]
                    )
                    steps_obj.set_fulfillment(
                        raw_obj.forms[i]["steps"][j]["fulfillment"]
                    )
                    steps_obj.set_event_error(
                        raw_obj.forms[i]["steps"][j]["event_error"]
                    )
                    steps_obj.set_personal_details(
                        raw_obj.forms[i]["steps"][j]["personal_details"]
                    )
                    steps_obj.set_summary(raw_obj.forms[i]["steps"][j]["summary"])
                    steps_obj.set_product_recommendation(
                        raw_obj.forms[i]["steps"][j]["product_recommendation"]
                    )
                    steps_obj.set_terms_and_condition(
                        raw_obj.forms[i]["steps"][j]["terms_and_condition"]
                    )
                    steps_obj.set_confirmation(
                        raw_obj.forms[i]["steps"][j]["confirmation"]
                    )
                    steps_obj.set_is_paperless(
                        raw_obj.forms[i]["steps"][j]["is_paperless"]
                    )
                    steps_obj.set_advertising(
                        raw_obj.forms[i]["steps"][j]["advertising"]
                    )
                    if raw_obj.forms[i]["steps"][j]["event_error"] == "true":
                        steps_obj.set_event_error("true")
                        steps_obj.set_error_message("{dynamic}")
                        steps_obj.set_error_code("{dynamic}")
                        steps_obj.set_error_type("{dynamic}")
                        steps_obj.set_error_sub_type("{dynamic}")
                        steps_obj.set_error_field("{dynamic}")
                    else:
                        steps_obj.set_event_error("false")

                    steps_obj.events_advertising = "true"
                    steps_obj.events_ad_click = "true"
                    steps_obj.advertising_tracking_code = "{dynamic}"
                    steps_obj.advertising_location = "{dynamic}"
                    steps_obj.advertising_type = "{dynamic}"
                    steps_obj.set_is_login(raw_obj.forms[i]["steps"][j]["is_login"])
                    steps_obj.set_is_otvc(raw_obj.forms[i]["steps"][j]["is_otvc"])
                    step_raw = json.dumps(steps_obj.__dict__)
                    step_raw_json = json.loads(step_raw)
                    temp_step.append(step_raw_json)
                    steps_prompt_obj = ""
                    steps_prompt_obj = steps_obj
                    steps_prompt_obj.set_is_otvc("")
                    str_prompt_temp = steps_obj.get_page_name + " (OTVC PROMPT)"
                    steps_prompt_obj.set_page_name(str(str_prompt_temp))
                    steps_prompt_obj.set_step_name(str(str_prompt_temp))
                    steps_prompt_obj.set_is_otvc_success_steps("")
                    steps_prompt_obj.set_page_view("false")
                    steps_prompt_obj.set_events_advertising("")
                    steps_prompt_obj.set_events_ad_click("")
                    steps_prompt_obj.set_advertising_location("")
                    steps_prompt_obj.set_advertising_tracking_code("")
                    steps_prompt_obj.set_advertising_type("")
                    steps_prompt_obj.set_is_otvc_prompt_steps("true")
                    steps_prompt_obj.set_user_id("{dynamic}")
                    steps_prompt_obj.set_user_type("{dynamic}")
                    steps_prompt_obj.set_form_view("")
                    steps_prompt_obj.set_form_qualify("")
                    steps_prompt_obj.set_form_submit("")
                    steps_prompt_obj.set_unique_id("")
                    steps_prompt_obj.set_to_transaction("")
                    steps_prompt_obj.set_from_transaction("")
                    steps_prompt_obj.set_login_event("")
                    steps_prompt_obj.set_step_name("")
                    steps_prompt_obj.set_personal_details("")
                    steps_prompt_obj.set_product_grouping("")
                    steps_prompt_obj.set_product_positioning("")
                    steps_prompt_obj.set_product_id("")
                    steps_prompt_obj.set_product_recommendation("")
                    steps_prompt_obj.set_parent_product("")
                    steps_prompt_obj.set_adjudication("")
                    steps_prompt_obj.set_is_joint("")
                    steps_prompt_obj.set_fulfillment("")
                    step = json.dumps(steps_prompt_obj.__dict__)
                    raw_prompt_steps = json.loads(step)
                    temp_step.append(raw_prompt_steps)
                    steps_success_obj = ""
                    steps_success_obj = steps_obj
                    steps_success_obj.set_is_otvc("")
                    str_success_temp = steps_obj.get_page_name
                    str_success_temp = str_success_temp.split(" (OTVC PROMPT)")
                    str_success_temp = str_success_temp[0] + " (OTVC SUCCESS)"
                    steps_success_obj.set_page_name(str(str_success_temp))
                    steps_success_obj.set_step_name(str(str_success_temp))
                    steps_success_obj.set_page_view("false")
                    steps_success_obj.set_is_otvc_success_steps("true")
                    steps_success_obj.set_is_otvc_prompt_steps("")
                    steps_success_obj.set_events_advertising("")
                    steps_success_obj.set_events_ad_click("")
                    steps_success_obj.set_advertising_location("")
                    steps_success_obj.set_advertising_tracking_code("")
                    steps_success_obj.set_advertising_type("")
                    steps_success_obj.set_user_id("{dynamic}")
                    steps_success_obj.set_user_type("{dynamic}")
                    steps_success_obj.set_form_view("")
                    steps_success_obj.set_form_qualify("")
                    steps_success_obj.set_form_submit("")
                    steps_success_obj.set_unique_id("")
                    steps_success_obj.set_to_transaction("")
                    steps_success_obj.set_from_transaction("")
                    steps_success_obj.set_login_event("")
                    steps_success_obj.set_step_name("")
                    steps_success_obj.set_personal_details("")
                    steps_success_obj.set_product_grouping("")
                    steps_success_obj.set_product_positioning("")
                    steps_success_obj.set_product_id("")
                    steps_success_obj.set_product_recommendation("")
                    steps_success_obj.set_parent_product("")
                    steps_success_obj.set_adjudication("")
                    steps_success_obj.set_is_joint("")
                    steps_success_obj.set_fulfillment("")
                    step = json.dumps(steps_success_obj.__dict__)
                    raw_success_steps = json.loads(step)
                    temp_step.append(raw_success_steps)
                else:
                    steps_obj = Steps()
                    steps_obj.set_step_name(raw_obj.forms[i]["steps"][j]["step_name"])
                    if data['application_type'] == APPLICATION_EMBER.lower():
                        page_hierarchy_steps_str = raw_obj.forms[i]["steps"][j]["page_hierarchy"]
                        page_name_steps_str = raw_obj.forms[i]["steps"][j]["page_name"]
                        page_name_steps_str = str(page_name_steps_str)
                        page_hierarchy_steps_str = str(page_hierarchy_steps_str)
                        steps_obj.set_page_hierarchy(page_hierarchy_steps_str)
                        steps_obj.set_page_name(page_name_steps_str)
                        steps_obj.set_page_path(page_hierarchy_steps_str)
                    if data['application_type'] == APPLICATION_ANGULAR.lower():
                        page_name_steps_str = raw_obj.forms[i]["steps"][j]["page_name"]
                        page_hierarchy_steps_str = raw_obj.forms[i]["steps"][j]["page_hierarchy"]
                        page_name_steps_str = str(page_name_steps_str)
                        page_hierarchy_steps_str = str(page_hierarchy_steps_str)
                        steps_obj.set_page_hierarchy(page_hierarchy_steps_str)
                        steps_obj.set_page_path(page_name_steps_str)
                        steps_obj.set_page_hierarchy(page_hierarchy_steps_str)
                    steps_obj.set_user_id(raw_obj.forms[i]["steps"][j]["user_id"])
                    steps_obj.set_user_type(raw_obj.forms[i]["steps"][j]["user_type"])
                    steps_obj.set_user_auth_state(
                        raw_obj.forms[i]["steps"][j]["user_auth_state"]
                    )
                    steps_obj.set_from_transaction(
                        raw_obj.forms[i]["steps"][j]["from_transaction"]
                    )
                    steps_obj.set_to_transaction(
                        raw_obj.forms[i]["steps"][j]["to_transaction"]
                    )
                    steps_obj.set_is_external(
                        raw_obj.forms[i]["steps"][j]["is_external"]
                    )
                    steps_obj.set_form_view(raw_obj.forms[i]["steps"][j]["form_view"])
                    steps_obj.set_form_qualify(
                        raw_obj.forms[i]["steps"][j]["form_qualify"]
                    )
                    steps_obj.set_form_submit(
                        raw_obj.forms[i]["steps"][j]["form_submit"]
                    )
                    steps_obj.set_product_id(raw_obj.forms[i]["steps"][j]["product_id"])
                    steps_obj.set_parent_product(
                        raw_obj.forms[i]["steps"][j]["parent_product"]
                    )
                    steps_obj.set_adjudication(
                        raw_obj.forms[i]["steps"][j]["adjudication"]
                    )
                    steps_obj.set_product_positioning(
                        raw_obj.forms[i]["steps"][j]["product_positioning"]
                    )
                    steps_obj.set_product_grouping(
                        raw_obj.forms[i]["steps"][j]["product_grouping"]
                    )
                    steps_obj.set_fulfillment(
                        raw_obj.forms[i]["steps"][j]["fulfillment"]
                    )
                    steps_obj.set_event_error(
                        raw_obj.forms[i]["steps"][j]["event_error"]
                    )
                    steps_obj.set_personal_details(
                        raw_obj.forms[i]["steps"][j]["personal_details"]
                    )
                    steps_obj.set_summary(raw_obj.forms[i]["steps"][j]["summary"])
                    steps_obj.set_product_recommendation(
                        raw_obj.forms[i]["steps"][j]["product_recommendation"]
                    )
                    steps_obj.set_terms_and_condition(
                        raw_obj.forms[i]["steps"][j]["terms_and_condition"]
                    )
                    steps_obj.set_confirmation(
                        raw_obj.forms[i]["steps"][j]["confirmation"]
                    )
                    steps_obj.set_is_paperless(
                        raw_obj.forms[i]["steps"][j]["is_paperless"]
                    )
                    steps_obj.set_advertising(
                        raw_obj.forms[i]["steps"][j]["advertising"]
                    )
                    if len(raw_obj.forms[i]["steps"][j]["from_transaction"]) != 0:
                        steps_obj.set_transaction_id("{dynamic}")
                    if len(raw_obj.forms[i]["steps"][j]["to_transaction"]) != 0:
                        steps_obj.set_transaction_id("{dynamic}")
                    steps_obj.events_advertising = "true"
                    steps_obj.events_ad_click = "true"
                    steps_obj.advertising_tracking_code = "{dynamic}"
                    steps_obj.advertising_location = "{dynamic}"
                    steps_obj.advertising_type = "{dynamic}"
                    steps_obj.set_is_login(raw_obj.forms[i]["steps"][j]["is_login"])
                    steps_obj.set_is_otvc(raw_obj.forms[i]["steps"][j]["is_otvc"])

                    if raw_obj.forms[i]["steps"][j]["event_error"] == "true":
                        steps_obj.set_event_error("true")
                        steps_obj.set_error_message("{dynamic}")
                        steps_obj.set_error_code("{dynamic}")
                        steps_obj.set_error_type("{dynamic}")
                        steps_obj.set_error_sub_type("{dynamic}")
                        steps_obj.set_error_field("{dynamic}")
                    else:
                        steps_obj.set_event_error("false")

                    steps_obj.set_is_transaction_complete("true")
                    step = json.dumps(steps_obj.__dict__)
                    raw_steps = json.loads(step)
                    temp_step.append(raw_steps)

                form_obj = {
                    "form_name": raw_obj.forms[i]["form_name"],
                    "no_of_steps": raw_obj.forms[i]["no_of_steps"],
                    "is_product_exist": raw_obj.forms[i]["is_product_exist"],
                    "is_transaction_exist": raw_obj.forms[i]["is_transaction_exist"],
                    "steps": temp_step,
                }
            temp_form_obj_dump = json.dumps(form_obj)
            temp_form_obj.append(json.loads(temp_form_obj_dump))
            temp_step = []

        raw_obj.forms.clear()
        raw_obj.forms = temp_form_obj
        json_page = json.dumps(raw_obj.__dict__)
        value_obj = json.loads(json_page)
        return value_obj

    def generate_VRG(self, data, raw_obj):
        req_data = self.get_otvc_generation(raw_obj, data)
        now = date.today()
        full = str(now.month) + "/" + str(now.day) + "/" + str(now.year)
        wb = openpyxl.Workbook()
        wb = load_workbook("VRG_template.xlsx")

        wb.template = True
        ws = wb.active
        ws.cell(row=6, column=2).value = "Last Modified:" + full
        ws.cell(row=6, column=4).value = (
            str(ws.cell(row=6, column=4).value) + str(req_data["user_name"]).title()
        )
        ws.cell(row=6, column=4).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        datestring = datetime.strftime(datetime.now(), " %Y_%m_%d %HH %MM")
        c = 8
        ws.cell(row=6, column=6).value = (
            str(ws.cell(row=6, column=6).value) + str(req_data["project_name"]).title()
        )
        fname = req_data["project_name"] + datestring

        for i in range(0, len(req_data["pages"])):
            r = 9
            if "page_name" in req_data["pages"][i]:
                ws.cell(row=8, column=c).value = str(
                    req_data["pages"][i]["page_name"]
                ).upper()
                ws.cell(row=8, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
            if "page_view" in req_data["pages"][i]:
                ws.cell(row=r, column=c).value = req_data["pages"][i]["page_view"]
            # site related stuff
            if "site_brand" in req_data:
                ws.cell(row=r + 1, column=c).value = req_data["site_brand"]
            if "site_name" in req_data:
                ws.cell(row=r + 2, column=c).value = req_data["site_name"]
            if "site_type" in req_data:
                ws.cell(row=r + 3, column=c).value = req_data["site_type"]
            if "site_environment" in req_data:
                ws.cell(row=r + 4, column=c).value = req_data["site_environment"]
            if "site_app_version" in req_data:
                ws.cell(row=r + 5, column=c).value = req_data["site_app_version"]
            if "site_last_build_date" in req_data:
                ws.cell(row=r + 6, column=c).value = req_data["site_last_build_date"]

            if "page_path" in req_data["pages"][i]:
                ws.cell(row=r + 7, column=c).value = req_data["pages"][i]["page_path"]

            if "page_name" in req_data["pages"][i]:
                ws.cell(row=r + 8, column=c).value = req_data["pages"][i]["page_name"]

            if "page_url" in req_data["pages"][i]:
                ws.cell(row=r + 9, column=c).value = req_data["pages"][i]["page_url"]
            if "page_referrer" in req_data["pages"][i]:
                ws.cell(row=r + 10, column=c).value = req_data["pages"][i][
                    "page_referrer"
                ]
            if "page_hierarchy" in req_data["pages"][i]:
                ws.cell(row=r + 11, column=c).value = str(
                    req_data["pages"][i]["page_hierarchy"]
                )
            if "page_language" in req_data["pages"][i]:
                ws.cell(row=r + 12, column=c).value = req_data["pages"][i][
                    "page_language"
                ]
            if "page_accessibility" in req_data["pages"][i]:
                ws.cell(row=r + 13, column=c).value = req_data["pages"][i][
                    "page_accessibility"
                ]

            if "user_auth_state" in req_data["pages"][i]:
                ws.cell(row=r + 14, column=c).value = req_data["pages"][i][
                    "user_auth_state"
                ]
            if "user_type" in req_data["pages"][i]:
                ws.cell(row=r + 15, column=c).value = req_data["pages"][i]["user_type"]
            if "user_id" in req_data["pages"][i]:
                ws.cell(row=r + 16, column=c).value = req_data["pages"][i]["user_id"]

            if "is_login" in req_data["pages"][i]:
                ws.cell(row=r + 42, column=c).value = req_data["pages"][i]["is_login"]

            if "event_error" in req_data["pages"][i]:
                ws.cell(row=r + 44, column=c).value = req_data["pages"][i][
                    "event_error"
                ]
            if "error_type" in req_data["pages"][i]:
                ws.cell(row=r + 45, column=c).value = req_data["pages"][i]["error_type"]
            if "error_sub_type" in req_data["pages"][i]:
                ws.cell(row=r + 46, column=c).value = req_data["pages"][i][
                    "error_sub_type"
                ]
            if "error_field" in req_data["pages"][i]:
                ws.cell(row=r + 47, column=c).value = req_data["pages"][i][
                    "error_field"
                ]
            if "error_message" in req_data["pages"][i]:
                ws.cell(row=r + 48, column=c).value = req_data["pages"][i][
                    "error_message"
                ]
            if "error_code" in req_data["pages"][i]:
                ws.cell(row=r + 49, column=c).value = req_data["pages"][i]["error_code"]
            if "events_advertising" in req_data["pages"][i]:
                ws.cell(row=r + 57, column=c).value = req_data["pages"][i][
                    "events_advertising"
                ]
            if "events_ad_click" in req_data["pages"][i]:
                ws.cell(row=r + 58, column=c).value = req_data["pages"][i][
                    "events_ad_click"
                ]
            if "advertising_tracking_code" in req_data["pages"][i]:
                ws.cell(row=r + 59, column=c).value = req_data["pages"][i][
                    "advertising_tracking_code"
                ]
            if "advertising_location" in req_data["pages"][i]:
                ws.cell(row=r + 60, column=c).value = req_data["pages"][i][
                    "advertising_location"
                ]
            if "advertising_type" in req_data["pages"][i]:
                ws.cell(row=r + 61, column=c).value = req_data["pages"][i][
                    "advertising_type"
                ]
            if "is_otvc_prompt" in req_data["pages"][i]:
                ws.cell(row=r + 96, column=c).value = req_data["pages"][i][
                    "is_otvc_prompt"
                ]
            if "is_otvc_success" in req_data["pages"][i]:
                ws.cell(row=r + 97, column=c).value = req_data["pages"][i][
                    "is_otvc_success"
                ]
            c = c + 1

        for parent in range(0, len(req_data["forms"])):
            r = 8
            for child in range(0, len(req_data["forms"][parent]["steps"])):
                if "page_name" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=8, column=c).value = (
                        str(req_data["forms"][parent]["form_name"])
                        .upper()
                        .replace(" ", "-")
                        + "|"
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        .upper()
                        .replace(" ", "-")
                    )
                    ws.cell(row=8, column=c).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                if "page_view" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 1, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_view"]

                if "site_brand" in req_data:
                    ws.cell(row=r + 2, column=c).value = req_data["site_brand"]
                if "site_name" in req_data:
                    ws.cell(row=r + 3, column=c).value = req_data["site_name"]
                if "site_type" in req_data:
                    ws.cell(row=r + 4, column=c).value = req_data["site_type"]
                if "site_environment" in req_data:
                    ws.cell(row=r + 5, column=c).value = req_data["site_environment"]
                if "site_app_version" in req_data:
                    ws.cell(row=r + 6, column=c).value = req_data["site_app_version"]
                if "site_last_build_date" in req_data:
                    ws.cell(row=r + 7, column=c).value = req_data[
                        "site_last_build_date"
                    ]
                if "page_path" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 8, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_path"]
                if "page_name" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 9, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_name"]

                if "page_url" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 10, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_url"]
                if "page_referrer" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 11, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_referrer"]
                if "page_hierarchy" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 12, column=c).value = str(
                        req_data["forms"][parent]["steps"][child]["page_hierarchy"]
                    )
                if "page_language" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 13, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_language"]
                if "page_accessibility" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 14, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_accessibility"]

                if "user_auth_state" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 15, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["user_auth_state"]
                if "user_type" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 16, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["user_type"]
                if "user_id" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 17, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["user_id"]

                if "form_view" in req_data["forms"][parent]["steps"][child]:
                    if req_data["forms"][parent]["steps"][child]["form_view"] == "true":
                        ws.cell(row=r + 18, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["form_view"]
                if "form_submit" in req_data["forms"][parent]["steps"][child]:
                    if (
                        req_data["forms"][parent]["steps"][child]["form_submit"]
                        == "true"
                    ):
                        ws.cell(row=r + 19, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["form_submit"]
                if "form_qualify" in req_data["forms"][parent]["steps"][child]:
                    if (
                        req_data["forms"][parent]["steps"][child]["form_qualify"]
                        == "true"
                    ):
                        ws.cell(row=r + 20, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["form_qualify"]
                if "form_name" in req_data["forms"][parent]:
                    ws.cell(row=r + 21, column=c).value = (
                        str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                    )
                if "unique_id" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 22, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["unique_id"]
                if "form_step" in req_data["forms"][parent]["steps"][child]:
                    if req_data["forms"][parent]["steps"][child]["form_step"] == "true":
                        ws.cell(row=r + 23, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["form_step"]

                if "step_name" in req_data["forms"][parent]["steps"][child]:
                    if req_data["forms"][parent]["steps"][child]["step_name"] != "":
                        ws.cell(row=r + 24, column=c).value = (
                            str(req_data["forms"][parent]["steps"][child]["step_name"])
                            .lower()
                            .replace(" ", "-")
                        )
                    else:
                        ws.cell(row=r + 24, column=c).value = self.get_page_hiearchy(
                            req_data["forms"][parent]["steps"][child]["page_hierarchy"]
                        )

                if "event_error" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 45, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["event_error"]
                if "error_type" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 46, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_type"]
                if "error_sub_type" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 47, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_sub_type"]
                if "error_field" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 48, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_field"]
                if "error_message" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 48, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_message"]
                if "error_code" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 50, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_code"]

                if req_data["forms"][parent]["is_transaction_exist"] == "true":
                    if "transaction_id" in req_data["forms"][parent]["steps"][child]:
                        ws.cell(row=r + 29, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["transaction_id"]
                    if (
                        "transaction_amount"
                        in req_data["forms"][parent]["steps"][child]
                    ):
                        ws.cell(row=r + 30, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["transaction_amount"]
                    if (
                        "transaction_service_fee"
                        in req_data["forms"][parent]["steps"][child]
                    ):
                        ws.cell(row=r + 31, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["transaction_service_fee"]
                    if "transaction_unit" in req_data["forms"][parent]["steps"][child]:
                        ws.cell(row=r + 32, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["transaction_unit"]
                    if "from_transaction" in req_data["forms"][parent]["steps"][child]:
                        ws.cell(row=r + 34, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["from_transaction"]
                    if "to_transaction" in req_data["forms"][parent]["steps"][child]:
                        ws.cell(row=r + 35, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["to_transaction"]
                    if "is_external" in req_data["forms"][parent]:
                        ws.cell(row=r + 36, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["is_external"]
                    if "is_transaction_complete" in req_data["forms"][parent]:
                        ws.cell(row=r + 37, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["is_transaction_complete"]

                # Product Specific Details
                if "product_id" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 87, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["product_id"]
                if "product_positioning" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 89, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["product_positioning"]
                if "product_grouping" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 90, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["product_grouping"]
                if "parent_product" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 91, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["parent_product"]
                if "fulfillment" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 81, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["fulfillment"]
                if "adjudication" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 93, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["adjudication"]
                if (
                    "product_recommendation"
                    in req_data["forms"][parent]["steps"][child]
                ):
                    if (
                        req_data["forms"][parent]["steps"][child][
                            "product_recommendation"
                        ]
                        == "true"
                    ):
                        ws.cell(row=r + 79, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["product_recommendation"]
                if "is_paperless" in req_data["forms"][parent]["steps"][child]:
                    if (
                        req_data["forms"][parent]["steps"][child]["is_paperless"]
                        == "true"
                    ):
                        ws.cell(row=r + 96, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["is_paperless"]
                if "summary" in req_data["forms"][parent]["steps"][child]:
                    if req_data["forms"][parent]["steps"][child]["summary"] == "true":
                        ws.cell(row=r + 80, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["summary"]
                if "confirmation" in req_data["forms"][parent]["steps"][child]:
                    if (
                        req_data["forms"][parent]["steps"][child]["confirmation"]
                        == "true"
                    ):
                        ws.cell(row=r + 83, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["confirmation"]
                if "terms_and_condition" in req_data["forms"][parent]["steps"][child]:
                    if (
                        req_data["forms"][parent]["steps"][child]["terms_and_condition"]
                        == "true"
                    ):
                        ws.cell(row=r + 82, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["terms_and_condition"]
                if "personal_details" in req_data["forms"][parent]["steps"][child]:
                    if (
                        req_data["forms"][parent]["steps"][child]["personal_details"]
                        == "true"
                    ):
                        ws.cell(row=r + 78, column=c).value = req_data["forms"][parent][
                            "steps"
                        ][child]["personal_details"]

                # Error Information in VRG
                if "event_error" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 45, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["event_error"]
                if "error_type" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 46, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_type"]
                if "error_sub_type" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 47, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_sub_type"]
                if "error_field" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 48, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_field"]
                if "error_message" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 49, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_message"]
                if "error_code" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 50, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["error_code"]
                if "is_login" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 43, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["is_login"]

                if "is_joint" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 58, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["is_joint"]
                if "events_advertising" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 58, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["events_advertising"]
                if "events_ad_click" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 59, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["events_ad_click"]
                if (
                    "advertising_tracking_code"
                    in req_data["forms"][parent]["steps"][child]
                ):
                    ws.cell(row=r + 60, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["advertising_tracking_code"]
                if "advertising_location" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 61, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["advertising_location"]
                if "advertising_type" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 62, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["advertising_type"]
                if "is_otvc_prompt_steps" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 97, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["is_otvc_prompt_steps"]
                if "is_otvc_success_steps" in req_data["forms"][parent]["steps"][child]:
                    ws.cell(row=r + 98, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["is_otvc_success_steps"]
                c = c + 1

        for i in range(0, len(req_data["interaction"])):
            r = 9
            if "interaction_name" in req_data["interaction"][i]:
                ws.cell(row=8, column=c).value = (
                    str(req_data["interaction"][i]["interaction_name"])
                    .upper()
                    .replace(" ", "-")
                    .replace(":", "-")
                )
                ws.cell(row=8, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
            if "page_view" in req_data["interaction"][i]:
                ws.cell(row=r, column=c).value = req_data["interaction"][i]["page_view"]

            # site related stuff
            if "site_brand" in req_data:
                ws.cell(row=r + 1, column=c).value = req_data["site_brand"]
            if "site_name" in req_data:
                ws.cell(row=r + 2, column=c).value = req_data["site_name"]
            if "site_type" in req_data:
                ws.cell(row=r + 3, column=c).value = req_data["site_type"]
            if "site_environment" in req_data:
                ws.cell(row=r + 4, column=c).value = req_data["site_environment"]
            if "site_app_version" in req_data:
                ws.cell(row=r + 5, column=c).value = req_data["site_app_version"]
            if "site_last_build_date" in req_data:
                ws.cell(row=r + 6, column=c).value = req_data["site_last_build_date"]

            if "page_path" in req_data["interaction"][i]:
                ws.cell(row=r + 7, column=c).value = req_data["interaction"][i][
                    "page_path"
                ]
            if "page_name" in req_data["interaction"][i]:
                ws.cell(row=r + 8, column=c).value = req_data["interaction"][i][
                    "page_name"
                ]
            if "page_url" in req_data["interaction"][i]:
                ws.cell(row=r + 9, column=c).value = req_data["interaction"][i][
                    "page_url"
                ]
            if "page_referrer" in req_data["interaction"][i]:
                ws.cell(row=r + 10, column=c).value = req_data["interaction"][i][
                    "page_referrer"
                ]
            if "page_hierarchy" in req_data["interaction"][i]:
                ws.cell(row=r + 11, column=c).value = req_data["interaction"][i][
                    "page_hierarchy"
                ]
            if "page_language" in req_data["interaction"][i]:
                ws.cell(row=r + 12, column=c).value = req_data["interaction"][i][
                    "page_language"
                ]
            if "page_accessibility" in req_data["interaction"][i]:
                ws.cell(row=r + 13, column=c).value = req_data["interaction"][i][
                    "page_accessibility"
                ]

            if "user_auth_state" in req_data["interaction"][i]:
                ws.cell(row=r + 14, column=c).value = req_data["interaction"][i][
                    "user_auth_state"
                ]
            if "user_type" in req_data["interaction"][i]:
                ws.cell(row=r + 15, column=c).value = req_data["interaction"][i][
                    "user_type"
                ]
            if "user_id" in req_data["interaction"][i]:
                ws.cell(row=r + 16, column=c).value = req_data["interaction"][i][
                    "user_id"
                ]

            if "site_interaction_event" in req_data["interaction"][i]:
                ws.cell(row=r + 55, column=c).value = req_data["interaction"][i][
                    "site_interaction_event"
                ]
            if "interaction_name" in req_data["interaction"][i]:
                ws.cell(row=r + 56, column=c).value = req_data["interaction"][i][
                    "interaction_name"
                ]
            c = c + 1

        for i in range(0, len(req_data["download"])):
            r = 9
            if "page_name" in req_data["download"][i]:
                ws.cell(row=8, column=c).value = (
                    str(req_data["download"][i]["download_file_name"])
                    .upper()
                    .replace(" ", "-")
                    .replace(":", "-")
                )
                ws.cell(row=8, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
            if "page_view" in req_data["download"][i]:
                ws.cell(row=r, column=c).value = req_data["download"][i]["page_view"]

            # site related stuff
            if "site_brand" in req_data:
                ws.cell(row=r + 1, column=c).value = req_data["site_brand"]
            if "site_name" in req_data:
                ws.cell(row=r + 2, column=c).value = req_data["site_name"]
            if "site_type" in req_data:
                ws.cell(row=r + 3, column=c).value = req_data["site_type"]
            if "site_environment" in req_data:
                ws.cell(row=r + 4, column=c).value = req_data["site_environment"]
            if "site_app_version" in req_data:
                ws.cell(row=r + 5, column=c).value = req_data["site_app_version"]
            if "site_last_build_date" in req_data:
                ws.cell(row=r + 6, column=c).value = req_data["site_last_build_date"]

            if "page_path" in req_data["download"][i]:
                ws.cell(row=r + 7, column=c).value = req_data["download"][i][
                    "page_path"
                ]
            if "page_name" in req_data["download"][i]:
                ws.cell(row=r + 8, column=c).value = req_data["download"][i][
                    "page_name"
                ]
            if "page_url" in req_data["download"][i]:
                ws.cell(row=r + 9, column=c).value = req_data["download"][i]["page_url"]
            if "page_referrer" in req_data["download"][i]:
                ws.cell(row=r + 10, column=c).value = req_data["download"][i][
                    "page_referrer"
                ]
            if "page_hierarchy" in req_data["download"][i]:
                ws.cell(row=r + 11, column=c).value = req_data["download"][i][
                    "page_hierarchy"
                ]
            if "page_language" in req_data["download"][i]:
                ws.cell(row=r + 12, column=c).value = req_data["download"][i][
                    "page_language"
                ]
            if "page_accessibility" in req_data["download"][i]:
                ws.cell(row=r + 13, column=c).value = req_data["download"][i][
                    "page_accessibility"
                ]

            if "user_auth_state" in req_data["download"][i]:
                ws.cell(row=r + 14, column=c).value = req_data["download"][i][
                    "user_auth_state"
                ]
            if "user_type" in req_data["download"][i]:
                ws.cell(row=r + 15, column=c).value = req_data["download"][i][
                    "user_type"
                ]
            if "user_id" in req_data["download"][i]:
                ws.cell(row=r + 16, column=c).value = req_data["download"][i]["user_id"]

            if "event_error" in req_data["download"][i]:
                ws.cell(row=r + 44, column=c).value = req_data["download"][i][
                    "event_error"
                ]

            if "events_download" in req_data["download"][i]:
                ws.cell(row=r + 50, column=c).value = req_data["download"][i][
                    "events_download"
                ]
            if "download_file_name" in req_data["download"][i]:
                ws.cell(row=r + 51, column=c).value = req_data["download"][i][
                    "download_file_name"
                ]
            if "download_file_type" in req_data["download"][i]:
                ws.cell(row=r + 52, column=c).value = req_data["download"][i][
                    "download_file_type"
                ]
            c = c + 1

        if req_data["mobile_json"] == "true":
            self.json_file(data, wb)
        if req_data["test_case"] == "true":
            self.test_case_vrg(data, wb)
            self.test_case_mobile_json(data, wb)

        wb.save(fname + ".xltx")
        return print("success")

    def json_file(self, req_data, wb):
        pages = {}
        for i in range(0, len(req_data["pages"])):
            if "application_type" in req_data:
                if "page_name" in req_data["pages"][i]:
                    pages[
                        "state_"
                        + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                        + ""
                    ] = {}
                    if (
                        req_data["pages"][i]["events_advertising"] == "true"
                        or req_data["pages"][i]["events_ad_click"] == "true"
                    ):
                        pages[
                            "state_"
                            + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                            + ""
                        ]["events"] = {}
                        if req_data["pages"][i]["events_ad_click"] == "true":
                            pages[
                                "state_"
                                + str(req_data["pages"][i]["page_name"]).replace(
                                    ".", "-"
                                )
                                + ""
                            ]["events"]["adclick"] = req_data["pages"][i][
                                "events_ad_click"
                            ]
                        if req_data["pages"][i]["events_advertising"] == "true":
                            pages[
                                "state_"
                                + str(req_data["pages"][i]["page_name"]).replace(
                                    ".", "-"
                                )
                                + ""
                            ]["events"]["adimpression"] = req_data["pages"][i][
                                "events_advertising"
                            ]
                    pages[
                        "state_"
                        + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                        + ""
                    ]["page"] = {}
                    pages[
                        "state_"
                        + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                        + ""
                    ]["page"]["name"] = self.get_page_name(
                        req_data["pages"][i]["page_hierarchy"]
                    )
                    pages[
                        "state_"
                        + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                        + ""
                    ]["page"]["hierarchy"] = self.get_page_hiearchy(
                        req_data["pages"][i]["page_hierarchy"]
                    )
                    if req_data["pages"][i]["events_advertising"] == "true":
                        pages[
                            "state_"
                            + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                            + ""
                        ]["advertising"] = {}
                        pages[
                            "state_"
                            + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                            + ""
                        ]["advertising"]["trackingcode"] = req_data["pages"][i][
                            "advertising_tracking_code"
                        ]
                        pages[
                            "state_"
                            + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                            + ""
                        ]["advertising"]["location"] = req_data["pages"][i][
                            "advertising_location"
                        ]
                        pages[
                            "state_"
                            + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                            + ""
                        ]["advertising"]["type"] = req_data["pages"][i][
                            "advertising_type"
                        ]
                    if (
                        req_data["pages"][i]["events_advertising"] == "true"
                        or req_data["pages"][i]["is_login"] == "true"
                    ):
                        pages[
                            "state_"
                            + str(req_data["pages"][i]["page_name"]).replace(".", "-")
                            + ""
                        ]["events"] = {}
                        if req_data["pages"][i]["events_advertising"] == "true":
                            pages[
                                "state_"
                                + str(req_data["pages"][i]["page_name"]).replace(
                                    ".", "-"
                                )
                                + ""
                            ]["events"]["adclick"] = req_data["pages"][i][
                                "events_ad_click"
                            ]
                            pages[
                                "state_"
                                + str(req_data["pages"][i]["page_name"]).replace(
                                    ".", "-"
                                )
                                + ""
                            ]["events"]["adimpression"] = req_data["pages"][i][
                                "events_advertising"
                            ]
                        if req_data["pages"][i]["is_login"] == "true":
                            pages[
                                "state_"
                                + str(req_data["pages"][i]["page_name"]).replace(
                                    ".", "-"
                                )
                                + ""
                            ]["events"]["login"] = req_data["pages"][i]["is_login"]
        interaction = {}
        for i in range(0, len(req_data["interaction"])):
            if "application_type" in req_data:
                if "interaction_name" in req_data["interaction"][i]:
                    interaction[
                        "action_"
                        + str(req_data["interaction"][i]["interaction_name"])
                        .lower()
                        .replace(" ", "-")
                        .replace(":", "-")
                        + ""
                    ] = {}
                    interaction[
                        "action_"
                        + str(req_data["interaction"][i]["interaction_name"])
                        .lower()
                        .replace(" ", "-")
                        .replace(":", "-")
                        + ""
                    ]["interaction"] = {}
                    interaction[
                        "action_"
                        + str(req_data["interaction"][i]["interaction_name"])
                        .lower()
                        .replace(" ", "-")
                        .replace(":", "-")
                        + ""
                    ]["interaction"]["name"] = (
                        str(req_data["interaction"][i]["interaction_name"])
                        .lower()
                        .replace(" ", "-")
                    )

        download = {}
        for i in range(0, len(req_data["download"])):
            download[
                "action_"
                + str(req_data["download"][i]["download_file_name"])
                .lower()
                .replace(" ", "-")
                .replace(":", "-")
                + ""
            ] = {}
            download[
                "action_"
                + str(req_data["download"][i]["download_file_name"])
                .lower()
                .replace(" ", "-")
                .replace(":", "-")
                + ""
            ]["events"] = {}
            download[
                "action_"
                + str(req_data["download"][i]["download_file_name"])
                .lower()
                .replace(" ", "-")
                .replace(":", "-")
                + ""
            ]["events"]["download"] = "true"
            download[
                "action_"
                + str(req_data["download"][i]["download_file_name"])
                .lower()
                .replace(" ", "-")
                .replace(":", "-")
                + ""
            ]["download"] = {}
            download[
                "action_"
                + str(req_data["download"][i]["download_file_name"])
                .lower()
                .replace(" ", "-")
                .replace(":", "-")
                + ""
            ]["download"]["filename"] = (
                str(req_data["download"][i]["download_file_name"])
                .lower()
                .replace(" ", "-")
            )
            download[
                "action_"
                + str(req_data["download"][i]["download_file_name"])
                .lower()
                .replace(" ", "-")
                .replace(":", "-")
                + ""
            ]["download"]["filetype"] = (
                str(req_data["download"][i]["download_file_type"])
                .lower()
                .replace(" ", "-")
            )

        forms = {}
        for parent in range(0, len(req_data["forms"])):
            for child in range(0, len(req_data["forms"][parent]["steps"])):
                if "page_name" in req_data["forms"][parent]["steps"][child]:
                    forms[
                        "state_"
                        + str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                        + "-"
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        .lower()
                        .replace(" ", "-")
                        + ""
                    ] = {}

                    # Product Object Created in Form JSON Object
                    if req_data["forms"][parent]["is_product_exist"] == "true":
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["product"] = [{}]
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["product"][0]["id"] = (
                            str(req_data["forms"][parent]["steps"][child]["product_id"])
                            .lower()
                            .replace(" ", "-")
                        )
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["product"][0]["positioning"] = (
                            str(
                                req_data["forms"][parent]["steps"][child][
                                    "product_positioning"
                                ]
                            )
                            .lower()
                            .replace(" ", "-")
                        )
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["product"][0]["grouping"] = (
                            str(
                                req_data["forms"][parent]["steps"][child][
                                    "product_grouping"
                                ]
                            )
                            .lower()
                            .replace(" ", "-")
                        )
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["product"][0]["parentproduct"] = str(
                            req_data["forms"][parent]["steps"][child]["parent_product"]
                        ).lower()
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["product"][0]["adjudication"] = str(
                            req_data["forms"][parent]["steps"][child]["adjudication"]
                        ).lower()
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["product"][0]["fulfilment"] = str(
                            req_data["forms"][parent]["steps"][child]["fulfillment"]
                        ).lower()
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["product"][0]["paperless"] = str(
                            req_data["forms"][parent]["steps"][child]["is_paperless"]
                        ).lower()

                    if req_data["forms"][parent]["is_transaction_exist"] == "true":
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["transaction"] = [{}]

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["transaction"][0]["id"] = (
                            str(
                                req_data["forms"][parent]["steps"][child][
                                    "transaction_id"
                                ]
                            )
                            .lower()
                            .replace(" ", "-")
                        )

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["transaction"][0]["items"] = [{}]

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["transaction"][0]["items"][0]["from"] = (
                            str(
                                req_data["forms"][parent]["steps"][child][
                                    "from_transaction"
                                ]
                            )
                            .lower()
                            .replace(" ", "-")
                        )

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["transaction"][0]["items"][0]["to"] = (
                            str(
                                req_data["forms"][parent]["steps"][child][
                                    "to_transaction"
                                ]
                            )
                            .lower()
                            .replace(" ", "-")
                        )

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["transaction"][0]["items"][0]["amount"] = (
                            str(
                                req_data["forms"][parent]["steps"][child][
                                    "transaction_amount"
                                ]
                            )
                            .lower()
                            .replace(" ", "-")
                        )

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["transaction"][0]["items"][0]["unit"] = (
                            str(
                                req_data["forms"][parent]["steps"][child][
                                    "transaction_unit"
                                ]
                            )
                            .lower()
                            .replace(" ", "-")
                        )

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["transaction"][0]["items"][0]["is_external"] = (
                            str(
                                req_data["forms"][parent]["steps"][child]["is_external"]
                            )
                            .lower()
                            .replace(" ", "-")
                        )

                    # Page Object Created in Form JSON Object
                    forms[
                        "state_"
                        + str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                        + "-"
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        .lower()
                        .replace(" ", "-")
                        + ""
                    ]["page"] = {}
                    forms[
                        "state_"
                        + str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                        + "-"
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        .lower()
                        .replace(" ", "-")
                        + ""
                    ]["page"]["hierarchy"] = self.get_page_hiearchy(
                        req_data["forms"][parent]["steps"][child]["page_hierarchy"]
                    )

                    if req_data["forms"][parent]["steps"][child]["step_name"] != "":
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["page"]["name"] = (
                            str(req_data["forms"][parent]["steps"][child]["step_name"])
                            .lower()
                            .replace(" ", "-")
                        )
                    else:
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["page"]["name"] = self.get_page_name(
                            req_data["forms"][parent]["steps"][child]["page_hierarchy"]
                        )

                    # Form Object Created in Form JSON Object
                    forms[
                        "state_"
                        + str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                        + "-"
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        .lower()
                        .replace(" ", "-")
                        + ""
                    ]["form"] = {}
                    forms[
                        "state_"
                        + str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                        + "-"
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        .lower()
                        .replace(" ", "-")
                        + ""
                    ]["form"]["name"] = (
                        str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                    )
                    if req_data["forms"][parent]["steps"][child]["step_name"] != "":
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["form"]["stepname"] = (
                            str(req_data["forms"][parent]["steps"][child]["step_name"])
                            .replace(" ", "-")
                            .replace(".", ">")
                        )
                    else:
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["form"]["stepname"] = self.get_page_name(
                            req_data["forms"][parent]["steps"][child]["page_hierarchy"]
                        )
                    if req_data["forms"][parent]["steps"][child]["step_name"] == "true":
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["advertising"] = {}

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["advertising"]["trackingcode"] = str(
                            req_data["forms"][parent]["steps"][child][
                                "advertising_tracking_code"
                            ]
                        ).lower()

                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["advertising"]["location"] = str(
                            req_data["forms"][parent]["steps"][child][
                                "advertising_location"
                            ]
                        ).lower()
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["advertising"]["type"] = str(
                            req_data["forms"][parent]["steps"][child][
                                "advertising_type"
                            ]
                        ).lower()

                    # Event Object created in Form JSON Object
                    forms[
                        "state_"
                        + str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                        + "-"
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        .lower()
                        .replace(" ", "-")
                        + ""
                    ]["events"] = {}
                    if (
                        req_data["forms"][parent]["steps"][child]["events_ad_click"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["adclick"] = str(
                            req_data["forms"][parent]["steps"][child]["events_ad_click"]
                        ).lower()
                    if (
                        req_data["forms"][parent]["steps"][child]["events_advertising"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["adimpression"] = str(
                            req_data["forms"][parent]["steps"][child][
                                "events_advertising"
                            ]
                        ).lower()
                    if req_data["forms"][parent]["steps"][child]["form_step"] == "true":
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["formstep"] = str(
                            req_data["forms"][parent]["steps"][child]["form_step"]
                        ).lower()
                    if req_data["forms"][parent]["steps"][child]["form_view"] == "true":
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["formview"] = str(
                            req_data["forms"][parent]["steps"][child]["form_view"]
                        ).lower()
                    if (
                        req_data["forms"][parent]["steps"][child]["form_qualify"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["formqualify"] = str(
                            req_data["forms"][parent]["steps"][child]["form_qualify"]
                        ).lower()
                    if (
                        req_data["forms"][parent]["steps"][child]["form_submit"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["formsubmit"] = str(
                            req_data["forms"][parent]["steps"][child]["form_submit"]
                        ).lower()

                    if (
                        req_data["forms"][parent]["steps"][child]["personal_details"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["applicationPersonalDetails"] = str(
                            req_data["forms"][parent]["steps"][child][
                                "personal_details"
                            ]
                        ).lower()
                    if (
                        req_data["forms"][parent]["steps"][child][
                            "product_recommendation"
                        ]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["applicationProductRecommendations"] = str(
                            req_data["forms"][parent]["steps"][child][
                                "product_recommendation"
                            ]
                        ).lower()
                    if req_data["forms"][parent]["steps"][child]["summary"] == "true":
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["applicationSummary"] = str(
                            req_data["forms"][parent]["steps"][child]["summary"]
                        ).lower()
                    if (
                        req_data["forms"][parent]["steps"][child]["confirmation"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["applicationConfirmation"] = str(
                            req_data["forms"][parent]["steps"][child]["confirmation"]
                        ).lower()
                    if (
                        req_data["forms"][parent]["steps"][child]["confirmation"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["applicationConfirmation"] = str(
                            req_data["forms"][parent]["steps"][child]["confirmation"]
                        ).lower()
                    if (
                        req_data["forms"][parent]["steps"][child]["terms_and_condition"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["applicationTermsConditions"] = str(
                            req_data["forms"][parent]["steps"][child][
                                "terms_and_condition"
                            ]
                        ).lower()
                    if (
                        req_data["forms"][parent]["steps"][child]["is_paperless"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["paperless"] = str(
                            req_data["forms"][parent]["steps"][child]["is_paperless"]
                        ).lower()

                    if (
                        req_data["forms"][parent]["steps"][child]["is_paperless"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["paperless"] = str(
                            req_data["forms"][parent]["steps"][child]["is_paperless"]
                        ).lower()

                    if (
                        req_data["forms"][parent]["steps"][child]["events_advertising"]
                        == "true"
                    ):
                        forms[
                            "state_"
                            + str(req_data["forms"][parent]["form_name"])
                            .lower()
                            .replace(" ", "-")
                            + "-"
                            + str(
                                req_data["forms"][parent]["steps"][child]["step_name"]
                            )
                            .lower()
                            .replace(" ", "-")
                            + ""
                        ]["events"]["login"] = str(
                            req_data["forms"][parent]["steps"][child]["is_login"]
                        ).lower()

        dictMerge = {**forms, **pages, **interaction, **download}

        with open("JsonData.json", "w") as outfile:
            json.dump(dictMerge, outfile)
        ws1 = wb.create_sheet("JSON File", 1)
        ws1.cell(row=1, column=1).value = json.dumps(dictMerge)
        ws1.row_dimensions[0].height = 300
        ws1.column_dimensions["B"].width = 300
        ws1.cell(row=1, column=1).font = Font(size=12, name="Times New Roman")

    def test_case_vrg(self, req_data, workbook):
        alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
        now = date.today()
        full = str(now.month) + "/" + str(now.day) + "/" + str(now.year)

        ws1 = workbook.create_sheet("VRG Test Cases", 2)
        ws1.row_dimensions[1].height = 34.5

        ws1.cell(row=1, column=1).value = "Sr. No."
        ws1.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        ws1.cell(row=1, column=2).value = "Date Created"
        ws1.cell(row=1, column=3).value = "Subject"
        ws1.cell(row=1, column=4).value = "Test Case Name"
        ws1.cell(row=1, column=5).value = "Test Objective"
        ws1.cell(row=1, column=6).value = "Step Name"
        ws1.cell(row=1, column=7).value = "Description"
        ws1.cell(row=1, column=8).value = "Expected Result"
        ws1.cell(row=1, column=9).value = "Author"
        for i in range(1, 10):
            ws1.cell(row=1, column=i).alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center", wrapText=True
            )
            ws1.cell(row=1, column=i).font = Font(bold=True, color="008000")

        ws1.column_dimensions["B"].width = 10
        ws1.column_dimensions["D"].width = 30
        ws1.column_dimensions["F"].width = 30
        ws1.column_dimensions["E"].width = 14.89
        ws1.column_dimensions["G"].width = 25
        ws1.column_dimensions["H"].width = 120

        id = 1
        r = 2
        c = 1
        s = "Step"
        e = (
            "To Verify Adobe variables match those detailed in the VRG \n \n "
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            "Pre Condition: Debugger tool to be launched for testing "
            "purpose(Recommended  Adobe Pulse Debugger)"
        )
        g = "Ensure Adobe Debugger is active"

        for i in range(0, len(req_data["pages"])):
            event = "events : "
            comment = " //"
            if "page_name" in req_data["pages"][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row=r, column=c + 1).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                ws1.cell(row=r, column=c + 2).value = str(
                    req_data["project_name"]
                ).title()
                ws1.cell(row=r, column=c + 2).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row=r, column=c + 4).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row=r, column=c + 6).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                val = req_data["pages"][i]["page_name"]
                b = "-."
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row=r, column=c + 5).value = req_data["pages"][i]["page_name"]

                ws1.cell(row=r, column=c + 3).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 5).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)

            if "site_brand" in req_data:
                if req_data["pages"][i]["events_advertising"] == "true":
                    event = event + " event76 ,"
                if req_data["pages"][i]["events_ad_click"] == "true":
                    event = event + "event77"
                if req_data["pages"][i]["is_login"] == "true":
                    event = event + "event19"
                if req_data["pages"][i]["is_otvc"] == "true":
                    event = event + ""
                dct = (
                    "page_name, eVar10: "
                    + str(req_data["site_brand"])
                    + ">"
                    + self.get_site_name_code(str(req_data["site_name"]))
                    + ">"
                    + str(req_data["pages"][i]["page_name"]).replace(".", ">")
                    + "    "
                    "//Unique for each Page"
                    + "\n eVar11: "
                    + str(req_data["site_app_version"])
                    + ":"
                    + str(req_data["site_last_build_date"])
                    + ":"
                    + self.get_site_name_code(str(req_data["site_name"]))
                    + ":"
                    + str(req_data["site_type"])
                    + "    //site_app_version,site_last_build_date,site_name,site_type"
                    + "\neVar15: "
                    + str(req_data["pages"][i]["page_url"])
                    + "    //Page URL"
                    + "\n eVar9: "
                    + str(req_data["pages"][i]["page_referrer"])
                    + "    //Previous Page Name"
                    + "\n eVar14: "
                    + str(req_data["pages"][i]["page_hierarchy"])
                    + "    // Hierarchy of the Page"
                    + "\n eVar16: "
                    + str(req_data["pages"][i]["page_language"])
                    + "    // Language Supported For "
                    "Page"
                    + "\n prop12: "
                    + self.get_page_accessibility_message(
                        str(req_data["pages"][i]["page_accessibility"])
                    )
                    + " //Page "
                    "Accessibility\n"
                    " "
                    "eVar19: {"
                    + str(req_data["pages"][i]["user_auth_state"])
                    + " : "
                    + str(req_data["pages"][i]["user_type"])
                    + "}"
                    + "    // Authentication of User "
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    "and User Type" + "\n eVar20: "
                    "" + str(req_data["pages"][i]["user_id"]) + "    // User ID\n"
                )
                dct = dct + event
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row=r, column=c + 7).alignment = alignment
                ws1.cell(row=r, column=c + 8).value = req_data["user_name"]
                ws1.cell(row=r, column=c + 8).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.row_dimensions[r].height = 180
            id = id + 1
            r = r + 1

        for i in range(0, len(req_data["interaction"])):
            event = "events :"
            comment = " "
            if "page_name" in req_data["interaction"][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row=r, column=c + 1).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )

                ws1.cell(row=r, column=c + 2).value = str(
                    req_data["project_name"]
                ).title()
                ws1.cell(row=r, column=c + 2).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row=r, column=c + 4).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row=r, column=c + 6).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                val = req_data["interaction"][i]["interaction_name"]
                b = "-."
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row=r, column=c + 3).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 5).value = req_data["interaction"][i][
                    "interaction_name"
                ]
                ws1.cell(row=r, column=c + 5).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)

            if "site_brand" in req_data:
                if bool(req_data["interaction"][i]["site_interaction_event"]):
                    event = event + " event24"
                    comment = comment + "// Site Interaction Event"
                    dct = (
                        "page_name, eVar10: "
                        + str(req_data["site_brand"])
                        + ">"
                        + self.get_site_name_code(str(req_data["site_name"]))
                        + ">"
                        + str(req_data["interaction"][i]["page_name"])
                        + "    //Unique for each Page"
                        + "\n eVar11: "
                        + str(req_data["site_app_version"])
                        + ":"
                        + str(req_data["site_last_build_date"])
                        + ":"
                        + self.get_site_name_code(str(req_data["site_name"]))
                        + ":"
                        + str(req_data["site_type"])
                        + "    //site_app_version,"
                        "site_last_build_date,"
                        "site_name,site_type"
                        + "\n eVar15: "
                        + str(req_data["interaction"][i]["page_url"])
                        + "    //Page URL"
                        + "\n eVar9: "
                        + str(req_data["interaction"][i]["page_referrer"])
                        + "    //Previous Page "
                        "Name"
                        + "\n eVar14: "
                        + str(req_data["interaction"][i]["page_hierarchy"])
                        + "    // Hierarchy of the "
                        "Page"
                        + "\n eVar16: "
                        + str(req_data["interaction"][i]["page_language"])
                        + "    // Language Supported "
                        "For "
                        "Page"
                        + "\n prop12: "
                        + self.get_page_accessibility_message(
                            str(req_data["interaction"][i]["page_accessibility"])
                        )
                        + "//Page "
                        "Accessibility \n  "
                        "eVar19: {{"
                        "not-authenticated "
                        "| authenticated} "
                        ": {dynamic}}    "
                        "// Authentication "
                        "of User and "
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        "User "
                        "Type \n  "
                        "eVar20: "
                        "{dynamic}   "
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        " // "
                        "User ID"
                        + "\n eVar24: "
                        + str(req_data["interaction"][i]["interaction_name"])
                        + " //Interaction Name \n"
                        + event
                        + "\t"
                        + comment
                    )
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row=r, column=c + 7).alignment = alignment
                ws1.cell(row=r, column=c + 8).value = req_data["user_name"]
                ws1.cell(row=r, column=c + 8).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.row_dimensions[r].height = 180
            id = id + 1
            r = r + 1

        for i in range(0, len(req_data["download"])):
            event = "events : "
            comment = ""
            if "page_name" in req_data["download"][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row=r, column=c + 1).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )

                ws1.cell(row=r, column=c + 2).value = str(
                    req_data["project_name"]
                ).title()
                ws1.cell(row=r, column=c + 2).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row=r, column=c + 4).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row=r, column=c + 6).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )

                val = req_data["download"][i]["download_file_name"]
                b = "-."
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row=r, column=c + 3).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 5).value = req_data["download"][i][
                    "download_file_name"
                ]
                ws1.cell(row=r, column=c + 5).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)

            if "site_brand" in req_data:
                if bool(req_data["download"][i]["events_download"]):
                    event = event + " event21"
                    comment = comment + "// Download Event, "
                dct = (
                    "page_name, eVar10: "
                    + str(req_data["site_brand"])
                    + ">"
                    + self.get_site_name_code(str(req_data["site_name"]))
                    + ">"
                    + str(req_data["download"][i]["page_name"])
                    + "    //Unique for each Page"
                    + "\n eVar11: "
                    + str(req_data["site_app_version"])
                    + ":"
                    + str(req_data["site_last_build_date"])
                    + ":"
                    + self.get_site_name_code(str(req_data["site_name"]))
                    + ":"
                    + str(req_data["site_type"])
                    + "    //site_app_version,"
                    "site_last_build_date,site_name,"
                    "site_type"
                    + "\n eVar15: "
                    + str(req_data["download"][i]["page_url"])
                    + "    //Page URL"
                    + "\n eVar9: "
                    + str(req_data["download"][i]["page_referrer"])
                    + "    //Previous Page Name"
                    + "\n eVar14: "
                    + str(req_data["download"][i]["page_hierarchy"])
                    + "    // Hierarchy of the Page"
                    + "\n eVar16: "
                    + str(req_data["download"][i]["page_language"])
                    + "    // Language Supported For "
                    "Page"
                    + "\n prop12: "
                    + str(
                        self.get_page_accessibility_message(
                            req_data["download"][i]["page_accessibility"]
                        )
                    )
                    + "//Page Accessibility "
                    "\n "
                    "eVar19: {{"
                    "not-authenticated | "
                    "authenticated} : {"
                    "dynamic}}    // "
                    "Authentication of "
                    "User and User Type "
                    "\n" + "eVar20: {"
                    "dynamic}    // "
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    ""
                    "User ID"
                    + "\n eVar21: "
                    + str(req_data["download"][i]["download_file_name"])
                    + ":"
                    + str(req_data["download"][i]["download_file_type"])
                    + "    // Download "
                    "FileName:FileType \n" + event + "\t" + comment
                )
            ws1.cell(row=r, column=c + 7).value = str(dct)
            ws1.cell(row=r, column=c + 7).alignment = openpyxl.styles.Alignment(
                horizontal="left", vertical="top", wrapText=True
            )

            ws1.cell(row=r, column=c + 8).value = req_data["user_name"]
            ws1.cell(row=r, column=c + 8).alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center", wrapText=True
            )

            ws1.row_dimensions[r].height = 180
            id = id + 1
            r = r + 1

        for parent in range(0, len(req_data["forms"])):
            for child in range(0, len(req_data["forms"][parent]["steps"])):
                event = "events : "
                comment = ""
                if req_data["forms"][parent]["steps"][child]["form_step"] == "true":
                    event = event + "event4"
                if req_data["forms"][parent]["steps"][child]["form_view"] == "true":
                    event = event + ", event1"
                if req_data["forms"][parent]["steps"][child]["form_submit"] == "true":
                    event = event + ", event2"
                if req_data["forms"][parent]["steps"][child]["form_qualify"] == "true":
                    event = event + ", event5"
                if req_data["forms"][parent]["steps"][child]["is_external"] == "true":
                    event = event + ", event46"
                if req_data["forms"][parent]["steps"][child]["is_login"] == "true":
                    event = event + ", event19"
                if req_data["forms"][parent]["steps"][child]["event_error"] == "true":
                    event = event + ", event3"
                if req_data["forms"][parent]["steps"][child]["is_paperless"] == "true":
                    event = event + ", event30"
                if (
                    req_data["forms"][parent]["steps"][child]["product_recommendation"]
                    == "true"
                ):
                    event = event + ", event103"
                if (
                    req_data["forms"][parent]["steps"][child]["events_advertising"]
                    == "true"
                ):
                    event = event + ", event77"
                if (
                    req_data["forms"][parent]["steps"][child]["events_ad_click"]
                    == "true"
                ):
                    event = event + ", event76"
                if req_data["forms"][parent]["steps"][child]["is_joint"] == "true":
                    event = event + ", event101"
                if "page_name" in req_data["forms"][parent]["steps"][child]:
                    ws1.cell(row=r, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_name"]
                    ws1.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )

                    ws1.cell(row=r, column=1).value = str(id)
                    ws1.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )

                    ws1.cell(row=r, column=2).value = full
                    ws1.cell(row=r, column=2).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )

                    ws1.cell(row=r, column=3).value = str(
                        req_data["project_name"]
                    ).title()
                    ws1.cell(row=r, column=3).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)

                    ws1.cell(row=r, column=5).value = e
                    ws1.cell(row=r, column=5).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )

                    ws1.cell(row=r, column=7).value = g
                    ws1.cell(row=r, column=7).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )

                    ws1.cell(row=r, column=4).value = str(
                        req_data["forms"][parent]["steps"][child]["page_name"]
                    ).replace(" ", "-")
                    ws1.cell(row=r, column=4).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=4).font = openpyxl.styles.Font(bold=True)
                    ws1.cell(row=r, column=6).value = str(
                        req_data["forms"][parent]["steps"][child]["page_name"]
                    ).replace(" ", "-")
                    ws1.cell(row=r, column=6).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=6).font = openpyxl.styles.Font(bold=True)
                    dt = (
                        "page_name, eVar10: "
                        + str(req_data["site_brand"])
                        + ">"
                        + self.get_site_name_code(str(req_data["site_name"]))
                        + ">"
                        + str(
                            req_data["forms"][parent]["steps"][child]["page_name"]
                        ).replace(".", ">")
                        + "    //Unique for "
                        "each Page \n "
                        "eVar11:"
                        + str(req_data["site_app_version"])
                        + ":"
                        + str(req_data["site_last_build_date"])
                        + ":"
                        + self.get_site_name_code(str(req_data["site_name"]))
                        + ":"
                        + str(req_data["site_type"])
                        + "    "
                        "//site_app_version:site_last_build_date:site_name"
                        ":site_type" + "\n eVar15: "
                        ""
                        + str(req_data["forms"][parent]["steps"][child]["page_url"])
                        + "    //Page URL"
                        + "\n eVar9: "
                        + str(
                            req_data["forms"][parent]["steps"][child]["page_referrer"]
                        )
                        + " //Previous Page Name"
                        + "\n eVar14: "
                        + str(
                            req_data["forms"][parent]["steps"][child]["page_hierarchy"]
                        )
                        + " // Hierarchy of the Page"
                        + "\n eVar16: "
                        + str(
                            req_data["forms"][parent]["steps"][child]["page_language"]
                        )
                        + "    // Language Supported "
                        "For Page"
                        + "\n prop12: "
                        + self.get_page_accessibility_message(
                            str(
                                req_data["forms"][parent]["steps"][child][
                                    "page_accessibility"
                                ]
                            )
                        )
                        + "//Page "
                        "Accessibility \n "
                        "eVar19: {"
                        + str(
                            req_data["forms"][parent]["steps"][child]["user_auth_state"]
                        )
                        + " : "
                        + str(req_data["forms"][parent]["steps"][child]["user_type"])
                        + "}"
                        + " // Authentication of "
                        "User and User Type "
                        + "\n eVar20: "
                        + str(req_data["forms"][parent]["steps"][child]["user_id"])
                        + " // User ID"
                        + "\n evar1: "
                        + str(req_data["forms"][parent]["form_name"])
                        .lower()
                        .replace(" ", "-")
                        .replace(".", "-")
                        + " // Form Name"
                        + "\n eVar4: "
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        + " //Form Step Name"
                        + " \n "
                        "eVar5: "
                        + str(req_data["forms"][parent]["steps"][child]["unique_id"])
                        + " // Unique Id of Form"
                    )
                    if req_data["forms"][parent]["is_transaction_exist"] == "true":
                        dt = (
                            dt
                            + "\n eVar41: "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "from_transaction"
                                ]
                            )
                            + ":"
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "to_transaction"
                                ]
                            )
                            + "    //Transaction "
                            "From: Transaction To"
                        )
                    if (req_data["forms"][parent]["is_product_exist"]) == "true":
                        dt = (
                            dt
                            + "\n eVar82: "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "product_positioning"
                                ]
                            )
                            + "    //Product "
                            "Positioning \n "
                            "eVar83 : "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "product_grouping"
                                ]
                            )
                            + "    //Product "
                            "grouping \n eVar84: "
                            ""
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "parent_product"
                                ]
                            )
                            + " // Parent Product "
                            "\neVar88:"
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "adjudication"
                                ]
                            )
                            + "   // Adjudication\n "
                        )
                    if (
                        req_data["forms"][parent]["steps"][child]["events_advertising"]
                    ) == "true":
                        dt = (
                            dt
                            + "advertising.trackingcode : "
                            + req_data["forms"][parent]["steps"][child][
                                "advertising_tracking_code"
                            ]
                            + " // "
                            "Advertising "
                            ""
                            ""
                            ""
                            "Tracking "
                            "Code \n "
                            "advertising.location :"
                            + req_data["forms"][parent]["steps"][child][
                                "advertising_location"
                            ]
                            + " // Advertising "
                            "Location \n "
                            "advertising.type: "
                            + req_data["forms"][parent]["steps"][child][
                                "advertising_type"
                            ]
                            + " // Advertsing Type\n"
                        )
                dt = dt + event
                ws1.cell(row=r, column=c + 7).value = str(dt)
                ws1.cell(row=r, column=c + 7).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )
                ws1.cell(row=r, column=c + 8).value = req_data["user_name"]
                ws1.cell(row=r, column=c + 8).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.row_dimensions[r].height = 203
                # ws1.row_dimensions[r].height = 180
                id = id + 1
                r = r + 1
            self.test_case_error(ws1, id, r, c, str(req_data["user_name"]))

    def test_case_mobile_json(self, req_data, workbook):
        border = Border(
            left=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
            top=Side(border_style="thin", color="FF000000"),
            bottom=Side(border_style="thin", color="FF000000"),
        )

        e = (
            "To Verify Adobe variables match those detailed in the JSON. \n "
            "\n Pre Condition: Charles Proxy tool to be launched for testing "
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            ""
            "purpose."
        )
        g = "Ensure Charles Proxy is active"

        note = (
            "Please note: All native requests are POST requests \n Check "
            "Request Headers in Charles Proxy"
        )
        developement = "Dev "
        production = "Prod "
        message = 'environment: All requests are for "POST ' "/b/ss/cibcnativeappdev/0/"
        code = "xxxxxx"

        ws1 = workbook.create_sheet("JSON Test Case", 3)
        ws1.row_dimensions[1].height = 24
        ws1.row_dimensions[2].height = 63.6

        # set the width of column in Mobile JSON Test Sheet
        ws1.column_dimensions["B"].width = 10
        ws1.column_dimensions["D"].width = 30
        ws1.column_dimensions["F"].width = 30
        ws1.column_dimensions["E"].width = 14.89
        ws1.column_dimensions["G"].width = 25
        ws1.column_dimensions["H"].width = 120

        # Column Header in Mobile JSON Test Case Sheet
        ws1.cell(row=1, column=1).value = "Sr. No."

        ws1.cell(row=1, column=2).value = "Date Created"
        ws1.cell(row=1, column=3).value = "Subject"
        ws1.cell(row=1, column=4).value = "Test Case Name"
        ws1.cell(row=1, column=5).value = "Test Objective"
        ws1.cell(row=1, column=6).value = "Step Name"
        ws1.cell(row=1, column=7).value = "Description"
        ws1.cell(row=1, column=8).value = "Expected Result"
        ws1.cell(row=1, column=9).value = "Author"

        for i in range(1, 10):
            ws1.cell(row=1, column=i).alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center", wrapText=True
            )
            ws1.cell(row=1, column=i).font = Font(bold=True, color="008000")
            ws1.cell(row=1, column=i).border = border

        ws1.cell(row=2, column=1).value = (
            note
            + "\n \t"
            + developement
            + message
            + code
            + "\n \t"
            + production
            + message
            + code
        )
        ws1.cell(row=2, column=1).alignment = openpyxl.styles.Alignment(
            horizontal="left", vertical="top", wrapText=True
        )
        ws1.cell(row=2, column=1).font = Font(color="008000")
        ws1.cell(row=2, column=1).border = border

        ws1.merge_cells("A2:G2")
        ws1.merge_cells("H2:Z2")

        id = 1
        r = 3
        c = 1
        now = date.today()
        full = str(now.month) + "/" + str(now.day) + "/" + str(now.year)

        for i in range(0, len(req_data["pages"])):
            event = "events: "
            comment = " //"
            ws1.row_dimensions[r].height = 203.6
            if "page_name" in req_data["pages"][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c).border = border

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row=r, column=c + 1).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 1).border = border

                ws1.cell(row=r, column=c + 2).value = str(
                    req_data["project_name"]
                ).title()
                ws1.cell(row=r, column=c + 2).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 2).border = border

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row=r, column=c + 4).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 4).border = border

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row=r, column=c + 6).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 6).border = border

                val = req_data["pages"][i]["page_name"]
                b = "-."
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row=r, column=c + 5).value = req_data["pages"][i]["page_name"]

                ws1.cell(row=r, column=c + 3).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 3).border = border
                ws1.cell(row=r, column=c + 5).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 5).border = border
            if "site_brand" in req_data:
                dct = (
                    "page_name: "
                    + str(req_data["site_brand"])
                    + ">"
                    + self.get_site_name_code(str(req_data["site_name"]))
                    + ">"
                    + str(req_data["pages"][i]["page_name"]).replace(".", ">")
                    + "  //Unique for each Page"
                    + "\n "
                    + "site.name : "
                    + self.get_site_name_code(str(req_data["site_name"]))
                    + "  //Name of the Site \n "
                    "site.type : "
                    + "mobile"
                    + "  //Type of the Site"
                    + "\n site.brand: "
                    + str(req_data["site_brand"])
                    + "// site brand \n "
                    "site.environment:"
                    + str(req_data["site_environment"])
                    + "\n page.previouspage: "
                    + str(req_data["pages"][i]["page_referrer"])
                    + "  //Previous Page Name"
                    + "\n page.hierarchy: "
                    + self.get_page_hiearchy(
                        req_data["pages"][i]["page_hierarchy"], "test"
                    )
                    + " // Hierarchy of the Page"
                    + ""
                    ""
                    + "\n page.accessibilty: "
                    + str(req_data["pages"][i]["page_accessibility"])
                    + " //Page Accessibility \n  "
                    "page.language : "
                    + str(req_data["pages"][i]["page_language"])
                    + "\n user.authstate:"
                    + str(req_data["pages"][i]["user_auth_state"])
                    + " // depending on pre sign on "
                    "or post sign on \n "
                    "user.type:"
                    + str(req_data["pages"][i]["user_type"])
                    + " //Type of the User \n user.id: "
                    + ""
                    + str(req_data["pages"][i]["user_id"])
                    + "   //user Id for the post sign on. it"
                    " "
                    ""
                    "should be consistent throughout the "
                    "whole app. " + "\n"
                )

                if req_data["pages"][i]["events_advertising"] == "true":
                    dct = (
                        dct
                        + "event.adImpression : "
                        + req_data["pages"][i]["events_advertising"]
                        + " // Ad Impression \n "
                        + "event.adclick : "
                        + req_data["pages"][i]["events_ad_click"]
                        + " //Ad Click \n "
                        "advertising.type : "
                        + req_data["pages"][i]["advertising_type"]
                        + "// Advertising Type \n advertising.location : "
                        + req_data["pages"][i]["advertising_location"]
                        + "// Advertising Location \n "
                        "advertising.trackingcode : "
                        + req_data["pages"][i]["advertising_tracking_code"]
                        + " // "
                        "Advertising "
                        "Tracking Code\n"
                    )
                if req_data["pages"][i]["is_login"] == "true":
                    dct = (
                        dct
                        + "events.login : "
                        + req_data["pages"][i]["is_login"]
                        + " //Login Event"
                    )

                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row=r, column=c + 7).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )
                ws1.cell(row=r, column=c + 7).border = border

                ws1.cell(row=r, column=c + 8).value = req_data["user_name"]
                ws1.cell(row=r, column=c + 8).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 8).border = border
                ws1.row_dimensions[r].height = 180
            id = id + 1
            r = r + 1
        for i in range(0, len(req_data["interaction"])):
            # ws1.row_dimensions[r].height = 203.6
            if "page_name" in req_data["interaction"][i]:

                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 2).border = border

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row=r, column=c + 1).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 1).border = border

                ws1.cell(row=r, column=c + 2).value = str(
                    req_data["project_name"]
                ).title()
                ws1.cell(row=r, column=c + 2).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 2).border = border

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row=r, column=c + 4).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )
                ws1.cell(row=r, column=c + 4).border = border

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row=r, column=c + 6).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 6).border = border

                val = req_data["interaction"][i]["interaction_name"]
                b = "-."
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row=r, column=c + 3).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 3).border = border

                ws1.cell(row=r, column=c + 5).value = req_data["interaction"][i][
                    "interaction_name"
                ]
                ws1.cell(row=r, column=c + 5).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 5).border = border

            dct = self.get_interaction_expected_result(i, req_data)
            if "site_brand" in req_data:
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row=r, column=c + 7).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )
                ws1.cell(row=r, column=c + 7).border = border

                ws1.cell(row=r, column=c + 8).value = req_data["user_name"]
                ws1.cell(row=r, column=c + 8).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 8).border = border

                ws1.row_dimensions[r].height = 203
            id = id + 1
            r = r + 1
        for parent in range(0, len(req_data["forms"])):
            for child in range(0, len(req_data["forms"][parent]["steps"])):
                if "page_name" in req_data["forms"][parent]["steps"][child]:
                    ws1.cell(row=r, column=c).value = req_data["forms"][parent][
                        "steps"
                    ][child]["page_name"]
                    ws1.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=1).value = str(id)
                    ws1.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=1).border = border

                    ws1.cell(row=r, column=2).value = full
                    ws1.cell(row=r, column=2).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )

                    ws1.cell(row=r, column=2).border = border

                    ws1.cell(row=r, column=3).value = str(
                        req_data["project_name"]
                    ).title()
                    ws1.cell(row=r, column=3).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)
                    ws1.cell(row=r, column=c + 3).border = border

                    ws1.cell(row=r, column=5).value = e
                    ws1.cell(row=r, column=5).alignment = openpyxl.styles.Alignment(
                        horizontal="left", vertical="top", wrapText=True
                    )
                    ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)
                    ws1.cell(row=r, column=c + 5).border = border

                    ws1.cell(row=r, column=7).value = g
                    ws1.cell(row=r, column=7).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=c + 7).border = border

                    val = req_data["forms"][parent]["steps"][child]["page_name"]

                    ws1.cell(row=r, column=4).value = str(val).title().replace(" ", "-")
                    ws1.cell(row=r, column=4).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=c + 4).border = border

                    ws1.cell(row=r, column=6).value = str(val).lower().replace(" ", "-")
                    ws1.cell(row=r, column=6).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=c + 6).border = border

                    dt = (
                        "page_name:"
                        + str(req_data["site_brand"])
                        + ">"
                        + self.get_site_name_code(str(req_data["site_name"]))
                        + ">"
                        + str(
                            req_data["forms"][parent]["steps"][child]["page_name"]
                        ).replace(".", ">")
                        + " //Unique for each"
                        " "
                        "Page \n "
                        + "site.name:"
                        + self.get_site_name_code(str(req_data["site_name"]))
                        + " //Site Name \n "
                        "site.type:" + "mobile" + "    "
                        "//Type of the Site"
                        + "\n site.brand: "
                        + str(req_data["site_type"])
                        + "  //Site Brand\n "
                        "site.environment: "
                        + str(req_data["site_environment"])
                        + "//Site Environment \n "
                        "page.previouspage: "
                        + str(
                            req_data["forms"][parent]["steps"][child]["page_referrer"]
                        )
                        + " //Previous Page Name"
                        + "\n page.hierarchy"
                        + self.get_page_hiearchy(
                            req_data["forms"][parent]["steps"][child]["page_hierarchy"],
                            "test",
                        )
                        + " // Hierarchy of the Page"
                        + "\n page.language: "
                        + str(
                            req_data["forms"][parent]["steps"][child]["page_language"]
                        )
                        + "// Language Supported For "
                        "Page" + "\n page.accessibility:"
                        " "
                        ""
                        ""
                        + str(
                            req_data["forms"][parent]["steps"][child][
                                "page_accessibility"
                            ]
                        )
                        + "//Page Accessibility "
                        "\n user.authstate:"
                        + str(
                            req_data["forms"][parent]["steps"][child]["user_auth_state"]
                        )
                        + "// depending on pre "
                        "sign "
                        "on or post sign on "
                        "\n "
                        "user.type"
                        + str(req_data["forms"][parent]["steps"][child]["user_type"])
                        + "}"
                        + " //Type of the User \n "
                        "user.id: "
                        + ""
                        + str(req_data["forms"][parent]["steps"][child]["user_id"])
                        + " //user Id for the post sign on. "
                        "it should be consistent "
                        "throughout the whole app. "
                        + "\n form.name: "
                        + str(req_data["forms"][parent]["form_name"])
                        + " // Form Name"
                        + "\n step.name: "
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        ""
                        + str(req_data["forms"][parent]["steps"][child]["step_name"])
                        + " //Form Step Name"
                        + " \n "
                        "unique.id: "
                        + str(req_data["forms"][parent]["steps"][child]["unique_id"])
                        + " // Unique Id of Form"
                    )
                    if req_data["forms"][parent]["is_transaction_exist"] == "true":
                        dt = (
                            dt
                            + "\n transaction.from: "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "from_transaction"
                                ]
                            )
                            + "// Transaction From "
                            "\n transaction.to: "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "to_transaction"
                                ]
                            )
                            + " //To Transaction"
                        )
                    if req_data["forms"][parent]["is_product_exist"] == "true":
                        dt = (
                            dt
                            + "\n positioning: "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "product_positioning"
                                ]
                            )
                            + "// Product "
                            "Positioning\n "
                            "grouping: "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "product_grouping"
                                ]
                            )
                            + "    //Product "
                            "grouping \n "
                            "parent_product: "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "parent_product"
                                ]
                            )
                            + " // Parent "
                            "Product "
                            "\nadjudication:"
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "adjudication"
                                ]
                            )
                            + "   // Adjudication"
                        )
                        if (
                            req_data["forms"][parent]["steps"][child][
                                "personal_details"
                            ]
                            == "true"
                        ):
                            dt = (
                                dt
                                + "\n application.personaldetails : "
                                + str(
                                    req_data["forms"][parent]["steps"][child][
                                        "personal_details"
                                    ]
                                )
                            )
                        if (
                            req_data["forms"][parent]["steps"][child]["summary"]
                            == "true"
                        ):
                            dt = (
                                dt
                                + "\n application.summary : "
                                + str(
                                    req_data["forms"][parent]["steps"][child]["summary"]
                                )
                            )
                        if (
                            req_data["forms"][parent]["steps"][child]["confirmation"]
                            == "true"
                        ):
                            dt = (
                                dt
                                + "\n application.confirmation : "
                                + str(
                                    req_data["forms"][parent]["steps"][child][
                                        "confirmation"
                                    ]
                                )
                            )
                        if (
                            req_data["forms"][parent]["steps"][child][
                                "product_recommendation"
                            ]
                            == "true"
                        ):
                            dt = (
                                dt + "\n application.productrecommendation "
                                ": "
                                + str(
                                    req_data["forms"][parent]["steps"][child][
                                        "product_recommendation"
                                    ]
                                )
                            )
                        if (
                            req_data["forms"][parent]["steps"][child][
                                "terms_and_condition"
                            ]
                            == "true"
                        ):
                            dt = (
                                dt
                                + "\n application.termsandcondition : "
                                + str(
                                    req_data["forms"][parent]["steps"][child][
                                        "terms_and_condition"
                                    ]
                                )
                            )
                        if (
                            req_data["forms"][parent]["steps"][child]["is_paperless"]
                            == "true"
                        ):
                            dt = (
                                dt
                                + "\n ispaperless : "
                                + str(
                                    req_data["forms"][parent]["steps"][child][
                                        "is_paperless"
                                    ]
                                )
                            )

                    if req_data["forms"][parent]["steps"][child]["form_view"] == "true":
                        dt = (
                            dt
                            + "\n event.formview: "
                            + str(
                                req_data["forms"][parent]["steps"][child]["form_view"]
                            )
                        )
                    if req_data["forms"][parent]["steps"][child]["form_step"] == "true":
                        dt = (
                            dt
                            + "\n event.formstep : "
                            + str(
                                req_data["forms"][parent]["steps"][child]["form_step"]
                            )
                        )
                    if (
                        req_data["forms"][parent]["steps"][child]["form_qualify"]
                        == "true"
                    ):
                        dt = (
                            dt
                            + "\n event.formqualify: "
                            + str(
                                req_data["forms"][parent]["steps"][child][
                                    "form_qualify"
                                ]
                            )
                        )
                    if (
                        req_data["forms"][parent]["steps"][child]["form_submit"]
                        == "true"
                    ):
                        dt = (
                            dt
                            + "\n event.submit : "
                            + str(
                                req_data["forms"][parent]["steps"][child]["form_submit"]
                            )
                        )
                    if (
                        req_data["forms"][parent]["steps"][child]["events_advertising"]
                    ) == "true":
                        dt = (
                            dt
                            + "\n advertising.trackingcode : "
                            + req_data["forms"][parent]["steps"][child][
                                "advertising_tracking_code"
                            ]
                            + " // "
                            "Advertising "
                            ""
                            ""
                            ""
                            "Tracking "
                            "Code \n "
                            "advertising.location :"
                            + req_data["forms"][parent]["steps"][child][
                                "advertising_location"
                            ]
                            + " // Advertising "
                            "Location \n "
                            "advertising.type: "
                            + req_data["forms"][parent]["steps"][child][
                                "advertising_type"
                            ]
                            + " // Advertsing "
                            "Type\n "
                            "events.adimpression "
                            ": "
                            + req_data["forms"][parent]["steps"][child][
                                "events_advertising"
                            ]
                            + " //Advertising "
                            "Impression \n "
                            "events.adclick: "
                            + req_data["forms"][parent]["steps"][child][
                                "events_ad_click"
                            ]
                            + " //Ad-Click Events"
                        )

                    if (
                        req_data["forms"][parent]["steps"][child]["is_login"]
                    ) == "true":
                        dt = (
                            dt
                            + "\n events.login : "
                            + req_data["forms"][parent]["steps"][child]["is_login"]
                            + " // Login Event"
                        )

                    ws1.cell(row=r, column=c + 7).value = str(dt)
                    ws1.cell(row=r, column=c + 7).alignment = openpyxl.styles.Alignment(
                        horizontal="left", vertical="top", wrapText=True
                    )
                    ws1.cell(row=r, column=c + 7).border = border

                    ws1.cell(row=r, column=c + 8).value = req_data["user_name"]
                    ws1.cell(row=r, column=c + 8).alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center", wrapText=True
                    )
                    ws1.cell(row=r, column=c + 8).border = border

                    ws1.row_dimensions[r].height = 203
                    id = id + 1
                    r = r + 1  # self.test_case_error(ws1, id, r,c,
                    # str(req_data['user_name']))
        for i in range(0, len(req_data["download"])):
            # ws1.row_dimensions[r].height = 203.6
            if "page_name" in req_data["download"][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 2).border = border

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row=r, column=c + 1).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 1).border = border

                ws1.cell(row=r, column=c + 2).value = str(
                    req_data["project_name"]
                ).title()
                ws1.cell(row=r, column=c + 2).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 2).border = border

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row=r, column=c + 4).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )
                ws1.cell(row=r, column=c + 4).border = border

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row=r, column=c + 6).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 6).border = border

                val = req_data["download"][i]["download_file_name"]
                ws1.cell(row=r, column=c + 3).value = str(val).title().replace(" ", "-")
                ws1.cell(row=r, column=c + 3).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 3).border = border

                ws1.cell(row=r, column=c + 5).value = str(val).lower().replace(" ", "-")
                ws1.cell(row=r, column=c + 5).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=r, column=c + 5).border = border

            dct = self.get_download_expected_result(dct, i, req_data)
            if "site_brand" in req_data:
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row=r, column=c + 7).alignment = openpyxl.styles.Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )
                ws1.cell(row=r, column=c + 7).border = border

                ws1.cell(row=r, column=c + 8).value = req_data["user_name"]
                ws1.cell(row=r, column=c + 8).alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
                ws1.cell(row=r, column=c + 8).border = border

                ws1.row_dimensions[r].height = 203
            id = id + 1
            r = r + 1

    def get_interaction_expected_result(self, i, req_data):
        interaction_test_case_expected_result = (
            "page_name"
            + str(req_data["site_brand"])
            + ">"
            + self.get_site_name_code(str(req_data["site_name"]))
            + ">"
            + str(req_data["interaction"][i]["page_name"])
            + "    //Unique for each Page {Interaction "
            "Page}"
            ""
            + "\n"
            + "site.name :"
            + self.get_site_name_code(str(req_data["site_name"]))
            + " //Name of the Site\n "
            "site.type: "
            "mobile" + "  //Type of the "
            "site"
            + "\n site.brand: "
            + str(req_data["site_brand"])
            + "//site brand \n site.environment : "
            + str(req_data["site_environment"])
            + "\n page.previouspage :"
            + str(req_data["interaction"][i]["page_referrer"])
            + "//Previous Page Name"
            + "\n page.hierarchy:"
            + self.get_page_hiearchy(
                req_data["interaction"][i]["page_hierarchy"], "test"
            )
            + "  // Hierarchy of the Page"
            + "\n page.language: "
            + str(req_data["interaction"][i]["page_language"])
            + "    // Language Supported "
            "For "
            "Page"
            + "\n page.accessibility:"
            + str(req_data["interaction"][i]["page_accessibility"])
            + "//Page Accessibility "
            + "\n interaction.name: "
            + str(req_data["interaction"][i]["interaction_name"])
            + " //Interaction Name \n "
            "event.siteInteraction: "
            + req_data["interaction"][i]["site_interaction_event"]
            + "// Site Interaction Event"
            + "\n user.id : {dynamic} "
            "//User "
            "Id" + "\n user.authstate : "
            "{"
            "not-authenticated | "
            "authenticated} // "
            "depending on pre "
            "sign on "
            ""
            ""
            "or post sign on "
        )

        return interaction_test_case_expected_result

    def get_download_expected_result(
        self, download_test_case_expected_result, i, req_data
    ):
        download_test_case_expected_result = (
            "page_name"
            + str(req_data["site_brand"])
            + ">"
            + self.get_site_name_code(str(req_data["site_name"]))
            + ">"
            + str(req_data["download"][i]["page_name"])
            + "    //Unique for each Page {Download Page}"
            ""
            + "\n"
            + "site.name :"
            + self.get_site_name_code(str(req_data["site_name"]))
            + " //Name of the Site\n site.type: "
            "mobile"
            + "  //Type of the site"
            + "\n site.brand: "
            + str(req_data["site_brand"])
            + "//site brand \n site.environment : "
            + str(req_data["site_environment"])
            + "\n page.previouspage :"
            + str(req_data["download"][i]["page_referrer"])
            + "//Previous Page Name"
            + "\n page.hierarchy:"
            + self.get_page_hiearchy(req_data["download"][i]["page_hierarchy"], "test")
            + "  // Hierarchy of the Page"
            + "\n page.language: "
            + str(req_data["download"][i]["page_language"])
            + "    // "
            ""
            ""
            "Language Supported For "
            "Page"
            + "\n page.accessibility:"
            + str(req_data["download"][i]["page_accessibility"])
            + "//Page Accessibility "
            + "\n file.name: "
            + str(req_data["download"][i]["download_file_name"])
            + " //File Name \n file.type: "
            + str(req_data["download"][i]["download_file_type"])
            + " // File Type"
            "event.download: " + req_data["download"][i]["events_download"] + "//  "
            "Download "
            "Event" + "\n user.id : {dynamic} //User "
            "Id" + "\n user.authstate : {"
            "not-authenticated | "
            "authenticated} // depending "
            "on "
            "pre sign on or post sign on "
        )

        return download_test_case_expected_result

    def test_case_error(self, worksheet, serial_no, row_number, column_no, user_name):
        now = date.today()
        full = str(now.month) + "/" + str(now.day) + "/" + str(now.year)
        worksheet.cell(row=row_number, column=column_no).value = str(serial_no)
        worksheet.cell(
            row=row_number, column=column_no
        ).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        worksheet.cell(row=row_number, column=column_no + 1).value = full
        worksheet.cell(
            row=row_number, column=column_no + 1
        ).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        worksheet.cell(row=row_number, column=column_no + 2).value = TEST_CASE_SUBJECT
        worksheet.cell(
            row=row_number, column=column_no + 2
        ).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.cell(
            row=row_number, column=column_no + 2
        ).font = openpyxl.styles.Font(bold=True)
        worksheet.cell(row=row_number, column=column_no + 4).value = TEST_CASE_OBJECTIVE
        worksheet.cell(
            row=row_number, column=column_no + 4
        ).alignment = openpyxl.styles.Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        worksheet.cell(
            row=row_number, column=column_no + 6
        ).value = TEST_CASE_ERROR_DESCRIPTION
        worksheet.cell(
            row=row_number, column=column_no + 6
        ).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.cell(
            row=row_number, column=column_no + 3
        ).value = TEST_CASE_ERROR_NAME
        worksheet.cell(
            row=row_number, column=column_no + 5
        ).value = TEST_CASE_ERROR_NAME
        worksheet.cell(
            row=row_number, column=column_no + 3
        ).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.cell(
            row=row_number, column=column_no + 3
        ).font = openpyxl.styles.Font(bold=True)
        worksheet.cell(
            row=row_number, column=column_no + 5
        ).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.cell(
            row=row_number, column=column_no + 5
        ).font = openpyxl.styles.Font(bold=True)
        worksheet.cell(
            row=row_number, column=column_no + 7
        ).value = TEST_CASE_ERROR_MESSAGE
        worksheet.cell(
            row=row_number, column=column_no + 7
        ).alignment = openpyxl.styles.Alignment(
            horizontal="left", vertical="top", wrapText=True
        )
        worksheet.cell(row=row_number, column=column_no + 8).value = user_name
        worksheet.cell(
            row=row_number, column=column_no + 8
        ).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
